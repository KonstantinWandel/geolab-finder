#!/usr/bin/env python
"""Append new source rows to `Geospatial_Data_Sources.xlsx`, idempotently.

Adding a source to the finder means adding a workbook row, because
`build_source_registry.py` derives the registry, the folder numbering and the record order
from that sheet. Doing it by hand is how a row ends up without its portal hyperlink (the
name cell carries it, invisible to pandas) or with a topic flag in the wrong column, and the
error only shows up as a source that never appears in a facet.

Rows written here are marked in AI blue (1F77B4), the workspace convention for machine-added
content in a human-authored document. Running twice adds nothing: a row whose name already
exists is skipped.

    python scripts/add_source_rows.py [--dry-run]
"""
from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Geospatial_Data_Sources.xlsx"
SHEET = "Tabelle1"
AI_BLUE = Font(color="FF1F77B4")

# One entry per source. `levels` and `topics` are header labels from row 1; they are matched
# case- and whitespace-insensitively, and an unknown label is a hard error rather than a
# silently unflagged column.
NEW_ROWS: List[Dict[str, Any]] = [
    {
        "name": "OpenStreetMap POI-Layer (Overpass)",
        "url": "https://wiki.openstreetmap.org/wiki/Map_features",
        "start_year": 2004,
        "end_year": 2026,
        "turnus": "laufend",
        "kommentar": "Overpass-API, freie Abfrage nach Tags; systematischer Ersatz für die "
                     "Spielplatz- und Arztsuchportale",
        "access": ["Direkter Download", "GUI", "API", "Karte"],
        "levels": ["Bundesland", "Kreise & kreisfreie Städte", "Gemeinden und Verbandsgemeinden",
                   "PLZ", "Adressen / Koordinaten", "weitere räumliche Gliederungen"],
        "topics": ["Bildung", "Anzahl Schulen", "Distanz nächste Schule/Hochschule/VHS",
                   "Gesundheit", "Distanz zur nächsten Apotheke",
                   "Distanz zum nächsten (Fach-)Arzt", "Kinder und Jugend",
                   "Distanz nächster Spielplatz", "Kultur", "Kinos",
                   "Distanz nächstes Theater", "Anzahl Gastronomiebetriebe / Bewohner",
                   "Religion", "Distanz zum nächsten Gotteshaus", "Sport",
                   "Sportanlagen / Einwohner", "Verkehr / Mobilität", "ÖPNV",
                   "Wirtschaft und Unternehmen", "Anzahl Läden / Bewohner"],
    },
    {
        "name": "Wegweiser Kommune (Bertelsmann Stiftung)",
        "url": "https://www.wegweiser-kommune.de/",
        "start_year": 2006,
        "end_year": 2040,
        "turnus": "jährlich",
        "kommentar": "Bevölkerungsprognosen bis 2040 für Gemeinden ab 5.000 Einwohnern",
        "access": ["Direkter Download", "GUI", "Karte"],
        "levels": ["Bundesland", "Kreise & kreisfreie Städte",
                   "Gemeinden und Verbandsgemeinden"],
        "topics": ["Bevölkerung", "Bevölkerungsstand", "Bevölkerung nach Geschlecht / Alter",
                   "Wanderungen", "Durchschnittsalter", "Arbeitsmarkt & Beschäftigung",
                   "Arbeitslose (-nquote)", "Bildung", "Betreuungsquote", "Soziales",
                   "Sozialleistungen", "Finanzen", "Schulden je Einwohner", "Bauen / Wohnen"],
    },
    {
        "name": "DWD Climate Data Center (CDC)",
        "url": "https://opendata.dwd.de/climate_environment/CDC/",
        "start_year": 1881,
        "end_year": 2026,
        "turnus": "täglich",
        "kommentar": "Stations- und Rasterdaten (1 km) zu Temperatur, Niederschlag, "
                     "Sonnenschein, Hitze- und Frosttagen; offen ohne Anmeldung",
        "access": ["Direkter Download", "API"],
        "levels": ["Bundesland", "Kreise & kreisfreie Städte", "Adressen / Koordinaten",
                   "weitere räumliche Gliederungen"],
        "topics": ["Nachhaltigkeit", "Umwelt", "Gesundheit"],
    },
    {
        "name": "BORIS-D Bodenrichtwerte",
        "url": "https://www.bodenrichtwerte-boris.de/",
        "start_year": 2000,
        "end_year": 2026,
        "turnus": "jährlich",
        "kommentar": "Bodenrichtwerte der Gutachterausschüsse, Bodenrichtwertzone; "
                     "Länderportale mit WMS/WFS, Nutzungsbedingungen je Land",
        "access": ["GUI", "API", "Karte"],
        "levels": ["Bundesland", "Kreise & kreisfreie Städte",
                   "Gemeinden und Verbandsgemeinden", "Adressen / Koordinaten",
                   "weitere räumliche Gliederungen"],
        "topics": ["Bauen / Wohnen", "Bautätigkeit und Wohnen", "Flächennutzung",
                   "Bodennutzung"],
    },
    {
        "name": "Wahlergebnisse Bundeswahlleiterin",
        "url": "https://www.bundeswahlleiterin.de/",
        "start_year": 1949,
        "end_year": 2025,
        "turnus": "je Wahl",
        "kommentar": "Ergebnisse und Strukturdaten für Bundestags- und Europawahlen, "
                     "Wahlkreis- und Gemeindeebene als CSV",
        "access": ["Direkter Download", "GUI"],
        "levels": ["Bundesland", "Kreise & kreisfreie Städte",
                   "Gemeinden und Verbandsgemeinden", "weitere räumliche Gliederungen"],
        "topics": ["Politik", "Bundestagswahl", "Europawahl", "Wahlberechtigte"],
    },
    {
        "name": "IÖR-Monitor (Flächennutzung)",
        "url": "https://www.ioer-monitor.de/",
        "start_year": 2000,
        "end_year": 2024,
        "turnus": "jährlich",
        "kommentar": "Rund 90 Indikatoren zur Flächennutzung und Landschaftsqualität, "
                     "Raster und Verwaltungsebenen, WMS/WFS",
        "access": ["GUI", "API", "Karte"],
        "levels": ["Bundesland", "Kreise & kreisfreie Städte",
                   "Gemeinden und Verbandsgemeinden", "weitere räumliche Gliederungen"],
        "topics": ["Flächennutzung", "nach ALKIS", "Bodennutzung", "Nachhaltigkeit", "Umwelt",
                   "Verkehr / Mobilität", "Straßenverkehr"],
    },
    {
        "name": "RWI-GEO-GRID / RWI-GEO-RED (FDZ Ruhr)",
        "url": "https://fdz.rwi-essen.de/",
        "start_year": 2005,
        "end_year": 2024,
        "turnus": "jährlich",
        "kommentar": "1-km-Raster mit sozioökonomischen Merkmalen und geocodierte "
                     "Immobilienanzeigen; Scientific-Use-Files auf Antrag",
        "access": ["Beantragung"],
        "levels": ["Bundesland", "Kreise & kreisfreie Städte",
                   "Gemeinden und Verbandsgemeinden", "PLZ", "Adressen / Koordinaten",
                   "weitere räumliche Gliederungen"],
        "topics": ["Bauen / Wohnen", "Mietpreise", "Wohnfläche je Wohnung / Einwohner",
                   "Bevölkerung", "Bevölkerungsstand", "Soziales", "Einkünfte",
                   "Armutsgefährdung"],
    },
]


def normalise(value: Any) -> str:
    return " ".join(str(value or "").split()).casefold()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(WORKBOOK)
    sheet = workbook[SHEET]
    header = {normalise(sheet.cell(1, column).value): column
              for column in range(1, sheet.max_column + 1)
              if sheet.cell(1, column).value}
    existing = {normalise(sheet.cell(row, 1).value)
                for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value}

    # the first fully empty row after the last named source
    next_row = 2
    for row in range(2, sheet.max_row + 2):
        if sheet.cell(row, 1).value:
            next_row = row + 1

    added: List[str] = []
    for entry in NEW_ROWS:
        if normalise(entry["name"]) in existing:
            print(f"[have] {entry['name']}")
            continue
        unknown = [label for label in entry["access"] + entry["levels"] + entry["topics"]
                   if normalise(label) not in header]
        if unknown:
            raise SystemExit(f"unknown workbook column(s) for {entry['name']}: {unknown}")

        cell = sheet.cell(next_row, 1, entry["name"])
        cell.hyperlink = entry["url"]
        cell.font = AI_BLUE
        for column, value in (("start year", entry.get("start_year")),
                              ("endyear", entry.get("end_year")),
                              ("Turnus", entry.get("turnus")),
                              ("Kommentar", entry.get("kommentar"))):
            if value is not None:
                target = sheet.cell(next_row, header[normalise(column)], value)
                target.font = AI_BLUE
        for label in entry["access"] + entry["levels"] + entry["topics"]:
            target = sheet.cell(next_row, header[normalise(label)], 1)
            target.font = AI_BLUE
        added.append(f"row {next_row}: {entry['name']}")
        next_row += 1

    if not added:
        print("nothing to add")
        return
    if args.dry_run:
        print("would add:\n  " + "\n  ".join(added))
        return
    backup = WORKBOOK.with_name(WORKBOOK.stem + "_prev_rows.xlsx")
    shutil.copy2(WORKBOOK, backup)
    workbook.save(WORKBOOK)
    print("added:\n  " + "\n  ".join(added))
    print(f"previous workbook kept as {backup.name}")
    print("next: bump EXPECTED_SOURCES in build_source_registry.py and re-run it")


if __name__ == "__main__":
    main()
