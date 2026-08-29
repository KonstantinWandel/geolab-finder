#!/usr/bin/env python3
"""Single source of truth for per-source progress.

Reads the registry, each source's `raw/` folder and `FETCH_LOG.json`, and the built
`geodb_metadata.json`, optionally re-checks every portal URL, and writes:

  data_sources/CHECKLIST.md          one line per source: state, artifacts, records, next step
  Geospatial_Data_Sources.xlsx       adds a machine-generated "Status_GeoDB" sheet
                                     (all AI-written cells in blue, per the workspace rule);
                                     the untouched original is kept as *_orig.xlsx

Run:
  python scripts/build_status_report.py                 # with a live link check
  python scripts/build_status_report.py --no-link-check
"""
from __future__ import annotations

import argparse
import json
import os
import http.cookiejar
import ssl
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_SOURCES = REPO_ROOT / "data_sources"
REGISTRY = DATA_SOURCES / "registry" / "geo_sources.json"
METADATA = REPO_ROOT / "soep_metadata_output" / "geodb_metadata.json"
WORKBOOK = REPO_ROOT / "Geospatial_Data_Sources.xlsx"
WORKBOOK_ORIGINAL = REPO_ROOT / "Geospatial_Data_Sources_orig.xlsx"
CHECKLIST = DATA_SOURCES / "CHECKLIST.md"
# Clean handoff folder: only what a human needs, named so it is identifiable on sight.
DELIVERABLES = REPO_ROOT / "deliverables_geodb_datenquellen"
PROGRESS_BASE = "geodb_datenquellen_fortschritt"

AI_BLUE = Font(color="FF1F77B4")
AI_BLUE_BOLD = Font(color="FF1F77B4", bold=True)

UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/122.0.0.0 Safari/537.36")

# Which builder source_key(s) each workbook source produces records under.
SOURCE_KEYS: Dict[str, List[str]] = {
    "regionalatlas-deutschland": ["regionalatlas"],
    "datenguide-abgeschaltet": ["regionalstatistik"],
    "strukturdaten-und-indikatoren-ba": ["ba_strukturdaten"],
    "strukturdaten-bundestagswahl-2021": ["btw_strukturdaten"],
    "migration-integration-in-regionen": ["migration_integration"],
    "hochschulkompass": ["hochschulkompass"],
    "laendermonitor-fruehkindliche-bildungssysteme": ["laendermonitor"],
    "deutschlandatlas-erreichbarkeit-von-apotheken": ["deutschlandatlas"],
    "krankenhausverzeichnis": ["gba_qualitaetsbericht"],
    "bundes-klinik-atlas": ["bundes_klinik_atlas"],
    "open-data-oepnv": ["opendata_oepnv", "transit_formats"],
    "german-companies": ["german_companies"],
    "unfallatlas": ["unfallatlas"],
    "arbeitsmarktreport-ba": ["ba_arbeitsmarktreport"],
    "deutsche-bahn-infrastrukturregister": ["db_isr"],
    "genesis-online-bund": ["genesis_bund"],
    "zensus-2022": ["zensus2022"],
    "breitband-monitor": ["breitband"],
    "arbeitsmarkt-kommunal-ba": ["ba_arbeitsmarkt_kommunal"],
    "arbeitsmarktstatistik-ba-karte": ["ba_glossar"],
    "open-data-handelsregister": ["offeneregister"],
    "destatis-regionale-mobilitaet-und-infektionsgesc": ["destatis_mobilitaet"],
    "deutsche-bahn-bahnhofsuche": ["db_stada"],
    "openstreetmap-poi-layer-overpass": ["osm_poi"],
    "wegweiser-kommune-bertelsmann-stiftung": ["wegweiser_kommune"],
    "dwd-climate-data-center-cdc": ["dwd_cdc"],
    "boris-d-bodenrichtwerte": ["boris_d"],
    "wahlergebnisse-bundeswahlleiterin": ["wahlergebnisse"],
    "ioer-monitor-flaechennutzung": ["ioer_monitor"],
    "rwi-geo-grid-rwi-geo-red-fdz-ruhr": ["fdz_ruhr"],
}

# Curated per-source state. `state` drives the checklist marker:
#   done     nothing outstanding
#   partial  indexed, but a better or fuller catalogue is still reachable
#   open     needs a human step before it can be indexed properly
# `next` is the developer detail (English, shown in CHECKLIST.md); `de` is the short phrase
# that goes into the German handoff table.
OPEN_ITEMS: Dict[str, Dict[str, str]] = {
    "regionalatlas-deutschland": {
        "de": "Vollständig eingebunden: alle 232 Indikatoren mit Beschreibung und direktem Link auf die passende Karte. Bei einer Aktualisierung des Atlas muss die Indikatorliste neu geladen werden.",
        "state": "done",
        "next": "Nothing outstanding. Re-fetch services.json when the atlas updates (it carries a timestamp per theme).",
    },
    "breitband-monitor": {
        "de": "Vollständig eingebunden: 632 Merkmale aus den Tabellen zu Breitband- und Mobilfunkversorgung, dazu die Struktur der beiden Rasterdatensätze (3,6 Mio. Gitterzellen). Die Rasterdaten selbst bleiben beim Anbieter, wir verweisen nur darauf.",
        "state": "done",
        "next": "632 indicators: the Breitbandatlas and Mobilfunk-Monitoring workbooks (use case x technology/bandwidth, Bund to Gemeinde) plus the two GeoPackages read at schema level (3.59 million grid cells x 168 coverage attributes, and 599,515 cells of mobile-operator counts). The GeoPackages themselves are never unpacked into the repo; scripts/extract_gpkg_schema.py writes a small schema JSON instead.",
    },
    "breitbandatlas": {
        "de": "Offen, inhaltlich aber abgedeckt: Die im Excel hinterlegte Adresse existiert nicht mehr. Der Nachfolger Gigabitgrundbuch ist als Zeile 2 vollständig eingebunden, diese Zeile ist damit doppelt.",
        "state": "open",
        "next": "The bmvi.de link in the workbook is dead; Gigabitgrundbuch is the successor. Needs the indicator/download page saved from a browser.",
    },
    "arbeitsmarktstatistik-ba-karte": {
        "de": "Vollständig eingebunden: 313 Begriffe aus dem amtlichen Glossar der Bundesagentur, jeweils mit deren eigener Definition. Die interaktive Karte veröffentlicht keine Liste ihrer Indikatoren; die Daten dahinter stecken in den Zeilen 5, 6 und 7.",
        "state": "done",
        "next": "313 concepts flattened from the BA Gesamtglossar, each with the BA's own definition, which is what a researcher searches for (Unterbeschaeftigung, Aktivierungsquote, Bedarfsgemeinschaft). The PDF is a two-column table whose column boundary moves between pages and whose long terms wrap, so it is parsed from real word coordinates (pdftotext -bbox-layout) with a lexical rule for wrapped terms; a fixed character split silently cut a quarter of the labels mid-word. The interactive map publishes no indicator catalogue of its own, and the BA data behind it is covered by the Arbeitsmarktreport, Strukturdaten and Arbeitsmarkt kommunal rows.",
    },
    "strukturdaten-und-indikatoren-ba": {
        "de": "Vollständig eingebunden: 68 Merkmale aus dem aktuellen Heft. Ein neueres Heft lässt sich jederzeit nachladen.",
        "state": "done",
        "next": "One booklet defines the series. Refresh with a newer heft when the BA publishes one.",
    },
    "arbeitsmarktreport-ba": {
        "de": "Vollständig eingebunden: 288 Merkmale aus 17 Tabellenblättern. Gleichnamige Merkmale tragen ihr Blatt im Titel, weil etwa 'Bestand an Arbeitslosen' je nach Blatt etwas anderes meint.",
        "state": "done",
        "next": "288 indicators flattened from the 17 data sheets (Eckwerte, SGB II/III, Unterbeschäftigung, Alo_Bestand/Bewegungen, Arbeitsstellen, Berufe, Ausbildung, Beschäftigung, Grundsicherung). Labels that recur across sheets carry their sheet in brackets, since 'Bestand an Arbeitslosen: Insgesamt' means something different in Eckwerte and in Eckwerte SGB II.",
    },
    "arbeitsmarkt-kommunal-ba": {
        "de": "Vollständig eingebunden: 33 gemeindescharfe Merkmale. Weitere Kreis-Hefte würden nur zusätzliche Regionen liefern, keine neuen Merkmale.",
        "state": "done",
        "next": "33 indicators flattened from one district archive (one XLSX per municipality, sheet 'Daten'). The indicator set is identical across districts, so more archives add regions, not concepts.",
    },
    "migration-integration-in-regionen": {
        "de": "Vollständig eingebunden: 140 Merkmale zu Migration und Integration in den Regionen.",
        "state": "done",
        "next": "Nothing outstanding.",
    },
    "krankenhausatlas-deutschland": {
        "de": "Offen und inhaltlich überholt: Die Daten stammen von 2016. Aktuelle Zahlen liefern die Zeilen 10 (Qualitätsberichte) und 11 (Klinik-Atlas), beide vollständig eingebunden.",
        "state": "open",
        "next": "Portal page only, and the atlas is at 2016. Superseded in practice by the G-BA Qualitätsberichte and the Bundes-Klinik-Atlas, both indexed.",
    },
    "krankenhausverzeichnis": {
        "de": "Vollständig eingebunden: 52 Abschnitte, die beschreiben, was ein Qualitätsbericht über ein Krankenhaus enthält. Die Berichte selbst (rund 1,7 GB je Jahr, einer je Haus) werden bewusst nicht ausgewertet: gesucht wird nach Themen, nicht nach einzelnen Häusern.",
        "state": "done",
        "next": "Schema sections indexed from the 2024 archive; the 2008-2024 archives are on disk (about 1.7 GB per year uncompressed, deliberately never extracted). Indexing the per-hospital rows would be a different product.",
    },
    "bundes-klinik-atlas": {
        "de": "Vollständig eingebunden: 41 Merkmale zu 1.577 Klinikstandorten mit Koordinaten. Ersetzt die eingestellte Weiße Liste.",
        "state": "done",
        "next": "Row renamed in the workbook on 2026-08-25: Weisse Liste is discontinued, the Bundes-Klinik-Atlas open-data export (IQTIG, 1,577 sites with coordinates) replaces it and is indexed.",
    },
    "arztsuche-bundesaerztekammer": {
        "de": "Offen, weil die Quelle nichts hergibt: Die Bundesärztekammer bietet nur eine Suchmaske, keinen Datenexport. Als systematischen Ersatz enthält der Finder die Arzt- und Zahnarztstandorte aus OpenStreetMap (Zeile 30).",
        "state": "open",
        "next": "Search UI over the Landesärztekammer registers, no export. Portal-level record only unless a state chamber publishes a list.",
    },
    "deutschlandatlas-erreichbarkeit-von-apotheken": {
        "de": "Vollständig eingebunden: alle 86 Indikatoren mit ihren amtlichen Definitionen, und alle 86 verlinken auf ihre eigene Kartenseite. Die Zeile deckt den gesamten Deutschlandatlas ab, nicht nur die Apothekenkarte.",
        "state": "done",
        "next": "All 86 indicators are indexed from the PDF and XLSX with their official definitions, and all 86 now link to their own map page (62 distinct maps out of the 122 the index lists). 83 are matched by title; the last three share no words with their map title and are pinned by Indikatorenkuerzel (bquali_mabschl, bquali_oabschl, v_5g), keyed on the code because the label carries the reference year. Correction to what was written here before: this host does NOT answer 400 to every scripted request. It answers 307 to a cookie-check URL and serves the page to any client that keeps the cookie, so the fetcher and both link checks use a cookie jar. A real map answers HTTP 200 with about 124 KB, a bogus code 404 with 96 KB.",
    },
    "hochschulkompass": {
        "de": "Vollständig eingebunden: die Merkmale des Hochschulregisters. Eine aktualisierte Liste lässt sich einfach austauschen.",
        "state": "done",
        "next": "Register attributes indexed. A refreshed hs_liste.txt is a drop-in replacement.",
    },
    "deutsche-bahn-infrastrukturregister": {
        "de": "Vollständig eingebunden: 415 Kartenebenen und Merkmale des Schienennetzes, unter anderem Elektrifizierung, Streckenklasse, Bahnsteige, Brücken und Bahnübergänge. Eine Anmeldung war entgegen der ursprünglichen Auskunft nicht nötig.",
        "state": "done",
        "next": "No registration needed, and it does publish a machine-readable catalogue after all. The viewer is a MapStore2 app over a public GeoServer: WMS GetCapabilities lists the map themes (Streckenklasse, Elektrifizierung, ETCS, Gleisanzahl, Betriebsstellen, Bahnsteige, Tunnel, Bruecken, Bahnuebergaenge) and WFS DescribeFeatureType lists the attributes per feature type. Both are indexed, so this row went from one portal card to 436 records. German and English field names are paired where ISR publishes both. Optional next step: DB's StaDa station dataset for the Bahnhofsuche row.",
    },
    "deutsche-bahn-bahnhofsuche": {
        "de": "Vollständig eingebunden: 37 Merkmale zu allen 5.408 Personenbahnhöfen, darunter Bahnhofskategorie, Barrierefreiheit, Ausstattung, zuständiger Aufgabenträger und Koordinaten. Möglich durch die von Ihnen besorgten Zugangsdaten.",
        "state": "done",
        "next": "Resolved 2026-08-27 with the DB API Marketplace credentials: the StaDa station API returns all 5,408 stations, and 37 attribute records plus one dataset record are indexed (Bahnhofskategorie, Barrierefreiheit, Mobilitaetsservice, Ausstattung, Aufgabentraeger, amtlicher Gemeindeschluessel, WGS84 coordinates), each with the measured coverage across stations. The station rows themselves are not indexed, following the register rule. Note the auth trap: the marketplace needs BOTH headers, DB-Client-Id and DB-Api-Key; the key alone answers 401 'Invalid client id or secret', which reads like a wrong key rather than a missing one.",
    },
    "open-data-oepnv": {
        "de": "Weitgehend eingebunden: 71 Datensätze und die Feldbeschreibungen der Fahrplanformate GTFS und NeTEx. Zum Herunterladen der Daten verlangt der Anbieter weiterhin ein kostenloses Konto; das Suchen und Finden funktioniert ohne.",
        "state": "done",
        "next": "71 named datasets indexed from the public catalogue (Deutschlandweite Sollfahrplandaten GTFS/NeTEX, Deutschlandweite Haltestellendaten, plus Soll-Fahrplandaten/Haltestellen/Liniendaten per Verbund), each with its own deep link, plus the no-login OpenService API products. Downloading a dataset still needs the free account.",
    },
    "spielplatztreff-suchmaschine-fuer-spielplaetze": {
        "de": "Offen, weil die Quelle nichts hergibt: eine Suchmaschine ohne Datenexport. Ersatz sind die Spielplätze aus OpenStreetMap (Zeile 30), bundesweit rund 136.000 Stück.",
        "state": "open",
        "next": "Crowd-sourced search UI, no export. Portal record only; OSM leisure=playground is the systematic alternative.",
    },
    "spielplatzkarte": {
        "de": "Offen wie Zeile 18: Kartenansicht ohne Datenexport, ersetzt durch die Spielplätze aus OpenStreetMap (Zeile 30).",
        "state": "open",
        "next": "Same as Spielplatztreff: map UI, no export.",
    },
    "destatis-regionale-mobilitaet-und-infektionsgesc": {
        "de": "Vollständig eingebunden, die Statistik selbst ist aber eingestellt: Sie lief nur von 2020 bis 2022. Die 6 Indikatorgruppen bleiben zitierfähig; auf jedem Eintrag steht, dass die Reihe beendet ist.",
        "state": "done",
        "next": "Discontinued experimental statistic (2020-2022), kept because the series stays citable. The six published indicator groups (Mobilitaetsindikatoren, zurueckgelegte Distanzen, Bewegungen nach Verkehrstraeger, Tagesverlauf) are flattened from the saved EXSTAT page and every record states that the statistic ended.",
    },
    "datenguide-abgeschaltet": {
        "de": "Vollständig eingebunden: der Merkmalskatalog der Regionaldatenbank (2.757 Merkmale) und 866 Regionaltabellen. Für 1.429 Merkmale wurde zusätzlich ermittelt, zu welcher Statistik sie gehören, sodass der Link nicht mehr auf der Startseite landet, sondern bei der richtigen Statistik.",
        "state": "done",
        "next": "Two catalogues under this row: the Datenguide GENESIS Merkmalskatalog (2,757 Merkmale) and the live Regionaldatenbank table catalogue (129 statistics, 866 tables, each with a working table-level deep link). The API reports exactly 129 statistics, so that enumeration is complete. On 2026-08-29 the weakest links in the whole index were fixed: 1,596 Merkmale pointed at the portal home page because no statistic code appears in their definition text. catalogue/statistics2variable answers that directly, so scripts/resolve_merkmal_statistics.py asked once per Merkmal and resolved 1,429 of 1,596 (89.5%); portal-level records here fell to 167 and statistic-level rose from 843 to 2,272. Federal statistic links are marked unverified because that portal is a client-rendered SPA that answers a 2.5 KB shell for any code, real or invented.",
    },
    "inkar": {
        "de": "Die Ursprungsquelle des Finders: 660 Indikatoren, unverändert enthalten.",
        "state": "done",
        "next": "Already the finder's original source (660 indicators).",
    },
    "german-companies": {
        "de": "Vollständig eingebunden: die Felder, die der Dienst je Unternehmen führt. Ein vollständiger Download ist nicht möglich, er beantwortet nur Anfragen zu einzelnen, namentlich bekannten Unternehmen.",
        "state": "done",
        "next": "API fields indexed and verified against live responses (samples in raw/). The endpoint is POST /lookup only, a record-linkage service: it resolves a company you already name. There is no bulk or search endpoint (every GET 404s, a city-only filter returns 0 rows), so the register cannot be downloaded through it, and the finder does not need it to. Key now lives in ~/kwandel/.config/secrets/.",
    },
    "open-data-handelsregister": {
        "de": "Vollständig eingebunden: 13 Felder, die das Handelsregister je Unternehmen führt, unter anderem Registernummer, Anschrift, Status, frühere Namen und Organe. Der vollständige Datenabzug ist mehrere Gigabyte groß und wird bewusst nicht gespeichert.",
        "state": "done",
        "next": "13 register fields indexed from the documented ocdata schema (Registernummer, Registergericht, Geschaeftsanschrift, Status, fruehere Firmennamen, Organe with entry and exit dates), so a query about company density, foundations, deletions or board networks routes here. The dump itself is multi-GB and is deliberately never downloaded: the address field is geocodable, which is what makes it a regional source.",
    },
    "laendermonitor-fruehkindliche-bildungssysteme": {
        "de": "Vollständig eingebunden: 17 Indikatoren mit den offiziellen Definitionen aus den Methodikpapieren. Die Indikatorliste der Website baut sich erst im Browser auf und ist maschinell nicht auslesbar; die Methodikpapiere sind die verbindliche Quelle.",
        "state": "done",
        "next": "All 17 indicators carry their official definition, extracted from the four public Methodik PDFs, which are the authoritative definition set. The portal's own indicator overview is a JavaScript app whose list is not in the HTML, so 17 is what is publicly parseable. Those PDFs are also two-column, so they only read correctly in reading order (pdftotext WITHOUT -layout); with -layout every definition picks up half a sentence from the neighbouring column.",
    },
    "genesis-online-bund": {
        "de": "Teilweise, und zwar wegen der Quelle, nicht wegen der Einbindung: Alle 3.026 Tabellen sind erfasst und direkt verlinkt, aber nur 55 davon reichen bis auf Kreis- oder Gemeindeebene, der Rest endet bei Bund und Ländern. Für regionale Fragen ist Zeile 21 die passendere Quelle.",
        "state": "partial",
        "next": "3,026 tables enumerated over the REST API (331 statistics) and indexed with table-level links, confirmed by hand in a browser on 2026-08-25. Mostly Bund/Land depth: only 55 titles name Kreise or Gemeinden, which is why this stays partial for a regional finder.",
    },
    "zensus-2022": {
        "de": "Teilweise, und zwar wegen der Quelle: Alle 1.440 Tabellen sind erfasst und verlinkt, für 1.407 ist die räumliche Ebene bestimmt. Nur 228 Tabellen sind feiner als Deutschland insgesamt, weil der Zensus viele Merkmale aus Datenschutzgründen nicht unterhalb der Länder veröffentlicht.",
        "state": "partial",
        "next": "1,440 tables from 12 statistics indexed with table-level links, confirmed by hand in a browser on 2026-08-25. The regional level is encoded in the opaque table code rather than the title; resolved per table through metadata/table on 2026-08-27: 1,407 of 1,440 tables now carry their real level, and a repeated title carries it in the label, which is what tells the four 'Personen: Religion' tables (Bundeslaender, Landeskirche, Bistum, Wahlkreise) apart. Remaining: nearly every table also has a national column, so 'Bund' appears alongside the finer level.",
    },
    "unfallatlas": {
        "de": "Vollständig eingebunden: 25 Merkmale der punktgenau erfassten Straßenverkehrsunfälle, Unfalljahre 2016 bis 2025. Die einzelnen Unfälle werden nicht indexiert.",
        "state": "done",
        "next": "Attributes of the geocoded accident records indexed (2016-2025, point level with WGS84 and UTM32 coordinates). The yearly CSV archives stay on disk; individual accidents are never indexed.",
    },
    "strukturdaten-bundestagswahl-2021": {
        "de": "Vollständig eingebunden: je 49 Merkmale für 2021 und 2025, getrennt geführt, weil sich Wahlkreiszuschnitt und Stichtage zwischen den beiden Wahlen unterscheiden. Die Wahlergebnisse selbst stehen in Zeile 34.",
        "state": "done",
        "next": "Both editions indexed: 49 indicators for the 2021 constituencies and 49 for 2025, kept separate because the same label is a different measurement under different constituency boundaries and reference dates. The indicator definitions come from the 2021 documentation page and carry over. Election RESULTS live in row 35 (Wahlergebnisse Bundeswahlleiterin).",
    },
    "openstreetmap-poi-layer-overpass": {
        "de": "Vollständig eingebunden: 26 Ortstypen wie Spielplätze, Apotheken, Arzt- und Zahnarztpraxen, Schulen, Kitas, Haltestellen, Sportanlagen und Ladesäulen, jeweils mit fertiger Abfrage und der gemessenen Anzahl in Deutschland.",
        "state": "done",
        "next": "26 POI layers indexed (Spielplaetze, Apotheken, Arztpraxen, Schulen, Kitas, Gotteshaeuser, Haltestellen, Sportanlagen, Ladesaeulen, ...), each with the Overpass query, a link to the tag's own wiki page and the measured object count for Germany. This is what closes rows 18/19 (playground portals) and complements row 12 (physician search), neither of which has an export. Counts come from taginfo.geofabrik.de, not Overpass: every Overpass mirror is refused at the network level from this host (connection refused, not a timeout), so a count query cannot run here even though the query in each record works elsewhere.",
    },
    "wegweiser-kommune-bertelsmann-stiftung": {
        "de": "Vollständig eingebunden: 393 Indikatoren mit Erläuterung, Berechnungsformel, Quelle und Zeitraum, darunter 67 Bevölkerungsprognosen bis 2040, die einzige Vorausberechnung in dieser Sammlung. Die Daten stehen unter einer freien Lizenz (CC0).",
        "state": "done",
        "next": "393 indicators indexed over the documented Data API, each with its explanation, calculation formula, source, unit, year range and finest region type, 67 of them population projections to 2040, which nothing else in this index has. This row was previously recorded as needing an account: it does not. /open-data documents an OpenAPI spec, the licence is CC0, and rest/indicator/list simply defaults to max=10, which is what made an unparameterised call look like a stub while the browse UI renders its tree client-side. A made-up friendly-url answers 404 on the API, so the identifiers are verified even though the human page is a SPA that renders 200 either way.",
    },
    "dwd-climate-data-center-cdc": {
        "de": "Vollständig eingebunden: 137 Klimaprodukte, unter anderem Temperatur, Niederschlag, Hitze-, Frost- und Sommertage, Sonnenschein und Bodenfeuchte, sowohl als deutschlandweites 1-km-Raster als auch als Messreihen der Wetterstationen. Frei und ohne Anmeldung herunterladbar.",
        "state": "done",
        "next": "137 records: one per (1-km grid or station) x aggregation x variable, read from the open Apache directory tree, which IS the DWD catalogue. Covers temperature, precipitation, sunshine, radiation, frost/hot/ice/summer days, snow cover, soil moisture, evaporation, wind, phenology and the vieljaehrige Mittel. Every record links to the directory that holds the files, and the grids are aggregable to Gemeinde or Kreis level, which is what makes climate joinable with the Regionalstatistik. Documentation, obsolete duplicates and project bundles are skipped rather than dressed up as indicators.",
    },
    "boris-d-bodenrichtwerte": {
        "de": "Weitgehend eingebunden: 5 Grundbegriffe und alle 16 Bundesländer. 14 Länder verweisen auf ihren eigenen Dienst; für Baden-Württemberg und Berlin führt das amtliche Verzeichnis keinen landesweiten Dienst, dort geht der Link auf das gemeinsame Portal BORIS-D.",
        "state": "done",
        "next": "Indexed as 5 concepts (Bodenrichtwert, Bodenrichtwertzone, Entwicklungs-/Beitragszustand, Nutzungsart, Immobilienrichtwerte) plus one record per Bundesland. BORIS-D is a viewer over the sixteen Laender services and publishes no catalogue of its own, so scripts/resolve_boris_services.py reads the services out of the official GDI-DE catalogue (1,962 Bodenrichtwerte records; per Land the best candidate is fetched in full and its GetCapabilities URL taken). 14 of 16 Laender now link to their own WMS/WFS or Bodenrichtwert portal, all probed HTTP 200; Baden-Wuerttemberg and Berlin publish no landesweiten service in that catalogue and keep the BORIS-D link rather than a guessed one. Three matching traps are written into the script: two-letter Land codes match inside unrelated titles, 'Sachsen' is a substring of both 'Niedersachsen' and 'Sachsen-Anhalt', and Niedersachsen and Bremen share one joint system.",
    },
    "wahlergebnisse-bundeswahlleiterin": {
        "de": "Vollständig eingebunden: 38 Merkmale und Datensätze, nämlich das Ergebnis der Bundestagswahl 2025, die Zeitreihe aller Bundestagswahlen ab 1949 und die Europawahl 2024, die bis auf Gemeindeebene reicht.",
        "state": "done",
        "next": "The result variables are read out of the 2025 kerg2 file itself: 4 system groups (Wahlberechtigte, Waehlende, Ungueltige, Gueltige) and one record per party that stood nationwide, plus dataset records for kerg/kerg2 2025, the results database since 1949, the Brief-/Urnenwahl series since 1957 and the Europawahl 2024 (which reaches Gemeinde level). Party lists that stood in a single Land are covered by the dataset records rather than one near-identical record each.",
    },
    "ioer-monitor-flaechennutzung": {
        "de": "Vollständig eingebunden: alle 88 Indikatoren zu Flächennutzung und Landschaftsqualität in 13 Kategorien. Die Karten- und Abrufdienste je Indikator verlangen ein kostenloses Nutzerkonto; die Indikatoren selbst sind frei einsehbar.",
        "state": "done",
        "next": "All 88 indicators indexed with their five-character code and category, read from the monitor's own public indicator list (linked from the 'Uebersicht der Geodienste' section of /indikatoren/). The earlier note here said the list sits behind the user area; that was wrong, only the SERVICE CALL needs a key. An unauthenticated monitor_api call answers a WMS ServiceException, so the records link to the indicator overview and carry the code plus the exact WMS/WFS/WCS call pattern instead of a per-indicator link that would not open for anyone.",
    },
    "rwi-geo-grid-rwi-geo-red-fdz-ruhr": {
        "de": "Vollständig eingebunden: 28 deutsche Datensätze mit ihrer Kennung (DOI), darunter das sozioökonomische 1-km-Raster und die geocodierten Immobilienanzeigen. Die Daten selbst gibt es nur auf Antrag beim Forschungsdatenzentrum, genau deshalb ist es nützlich, sie hier überhaupt zu finden.",
        "state": "done",
        "next": "28 German datasets indexed with title, DOI, keywords and abstract, harvested from the da|ra detail pages because the portal lists every dataset only as a 'Details' link. Includes RWI-GEO-GRID (1-km socio-economic grid), RWI-GEO-RED (geocoded real-estate advertisements), RWI-GEO-KITA, the regional house-price indices and the PLZ-to-Gemeinde bridge. 7 non-German evaluation studies (Burkina Faso, Senegal, Rwanda, India) are skipped as out of scope. These are scientific-use files on application, which is exactly why they belong in a finder: the point is to learn the dataset exists before starting an application.",
    },
}


# Sources NOT in the workbook that a German regional-data finder arguably should carry.
# Ordered by what they would add that nothing already indexed provides.
CANDIDATES = [
    ("OpenStreetMap / Overpass POI layers", "https://overpass-turbo.eu/",
     "The systematic replacement for the crowd-sourced portals in the workbook: playgrounds, "
     "pharmacies, GP practices, schools, kindergartens, stops, supermarkets, all as coordinates "
     "with a documented tag schema. Free, no registration, reproducible queries."),
    ("Wegweiser Kommune (Bertelsmann Stiftung)", "https://www.wegweiser-kommune.de/",
     "About 100 indicators for every municipality above 5,000 inhabitants plus demographic "
     "projections to 2040. Complements INKAR on the projection side, which nothing here has."),
    ("BORIS-D / Bodenrichtwerte", "https://www.bodenrichtwerte-boris.de/",
     "Official land values from the Gutachterausschüsse, parcel level. The land-price counterpart "
     "to INKAR's asking rents."),
    ("DWD Climate Data Center", "https://opendata.dwd.de/climate_environment/CDC/",
     "Station and gridded climate series (temperature, precipitation, heat days) at 1 km. The only "
     "environmental/climate axis; free and openly downloadable."),
    ("RWI-GEO-GRID / RWI-GEO-RED (FDZ Ruhr)", "https://fdz.rwi-essen.de/",
     "1 km grid socio-economic data and geocoded real-estate advertisements. Scientific-use files "
     "on application, heavily used in German regional research."),
    ("Election results (Bundeswahlleiter and the Länder)", "https://www.bundeswahlleiter.de/",
     "We index the 2021 structural data but not the results. Constituency and municipality level "
     "results for federal, European and state elections are downloadable as CSV."),
    ("IÖR-Monitor", "https://www.ioer-monitor.de/",
     "Around 90 land-use and landscape-quality indicators at fine spatial resolution, with a WMS/WFS "
     "API. Deeper on land use than the ALKIS shares in INKAR and Regionalatlas."),
]

# How precisely a record's outward link lands on the thing it describes.
LINK_LEVEL_WORD = {
    "indicator": "straight to the indicator",
    "table": "straight to the table",
    "statistic": "to the statistic containing it",
    "dataset": "to the dataset containing it",
    "portal": "to the portal (search from there)",
}

STATE_MARK = {"done": "[x]", "partial": "[~]", "open": "[ ]"}
STATE_WORD = {"done": "done", "partial": "partial", "open": "open"}


def check_url(url: str) -> str:
    if not url:
        return "no url"
    request = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
    }, method="GET")
    try:
        # deutschlandatlas.bund.de answers 307 to a cookie-check URL and only serves the page to
        # a client that keeps the cookie, so a jar is the difference between "unreachable" and
        # HTTP 200. Certificate hygiene is not what this check is about (inkar.de ships an
        # incomplete chain), so verification is off here as in the link auditor.
        opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(http.cookiejar.CookieJar()),
            urllib.request.HTTPSHandler(context=ssl._create_unverified_context()),
        )
        with opener.open(request, timeout=25) as response:
            return f"HTTP {response.status}"
    except urllib.error.HTTPError as exc:
        return f"HTTP {exc.code}"
    except (urllib.error.URLError, ssl.SSLError, OSError) as exc:
        reason = getattr(exc, "reason", exc)
        return f"unreachable ({str(reason)[:40]})"


def gather(link_check: bool) -> List[Dict[str, Any]]:
    registry = json.loads(REGISTRY.read_text(encoding="utf-8"))["sources"]
    records: List[Dict[str, Any]] = []
    if METADATA.exists():
        records = json.loads(METADATA.read_text(encoding="utf-8"))
    counts: Dict[str, int] = {}
    link_levels: Dict[str, Dict[str, int]] = {}
    unverified: Dict[str, int] = {}
    for record in records:
        key = record["source_key"]
        counts[key] = counts.get(key, 0) + 1
        level = record.get("link_level", "portal")
        link_levels.setdefault(key, {})
        link_levels[key][level] = link_levels[key].get(level, 0) + 1
        if record.get("link_verified") is False:
            unverified[key] = unverified.get(key, 0) + 1

    rows: List[Dict[str, Any]] = []
    for position, source in enumerate(registry, start=1):
        folder = DATA_SOURCES / f"{position:02d}-{source['slug']}"
        raw = folder / "raw"
        files = sorted(
            (p for p in raw.iterdir() if p.name not in {".gitkeep", "FETCH_LOG.json"}),
            key=lambda p: p.name,
        ) if raw.exists() else []
        size = sum(p.stat().st_size for p in files if p.is_file())
        log_path = raw / "FETCH_LOG.json"
        log = json.loads(log_path.read_text(encoding="utf-8")) if log_path.exists() else {"artifacts": {}}
        indexed = sum(counts.get(key, 0) for key in SOURCE_KEYS.get(source["slug"], []))
        levels: Dict[str, int] = {}
        unverified_count = 0
        for key in SOURCE_KEYS.get(source["slug"], []):
            for level, count in link_levels.get(key, {}).items():
                levels[level] = levels.get(level, 0) + count
            unverified_count += unverified.get(key, 0)
        item = OPEN_ITEMS.get(source["slug"], {"state": "open", "next": "not reviewed yet", "de": "noch nicht geprüft"})
        rows.append({
            "position": position,
            "name": source["name"],
            "slug": source["slug"],
            "url": source["url"],
            "folder": folder.name,
            "files": [p.name for p in files],
            "bytes": size,
            "fetched": len(log.get("artifacts", {})),
            "indexed": indexed,
            "portal_record": 1 if source["slug"] != "inkar" else 0,
            "link_levels": levels,
            "unverified_links": unverified_count,
            "state": item["state"],
            "next": item["next"],
            "next_de": item.get("de", item["next"]),
            "note": source["note"],
            "coverage": f"{source['coverage_start_year'] or '?'}-{source['coverage_end_year'] or '?'}"
                        if (source["coverage_start_year"] or source["coverage_end_year"]) else "",
            "access": ", ".join(source["access_modes"]),
            "link": "",
        })

    if link_check:
        with ThreadPoolExecutor(max_workers=8) as pool:
            results = list(pool.map(lambda row: check_url(row["url"]), rows))
        for row, result in zip(rows, results):
            row["link"] = result
    return rows


def human_bytes(size: int) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if size < 1024 or unit == "GB":
            return f"{size:.0f} {unit}" if unit == "B" else f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{size:.1f} GB"


def write_checklist(rows: List[Dict[str, Any]], stamp: str) -> None:
    total_indexed = sum(row["indexed"] for row in rows)
    done = sum(1 for row in rows if row["state"] == "done")
    partial = sum(1 for row in rows if row["state"] == "partial")
    open_count = sum(1 for row in rows if row["state"] == "open")

    lines: List[str] = []
    lines.append("# Source checklist")
    lines.append("")
    lines.append(f"Generated by `scripts/build_status_report.py` on {stamp}. Do not hand-edit: "
                 "change `OPEN_ITEMS` in that script and re-run.")
    lines.append("")
    lines.append(f"**{done} done, {partial} partial, {open_count} open** of {len(rows)} sources. "
                 f"{total_indexed} indicator-level records built (plus one portal-level record per source).")
    lines.append("")
    lines.append("`[x]` nothing outstanding | `[~]` indexed, a fuller catalogue is still reachable | "
                 "`[ ]` needs a human step")
    lines.append("")
    totals: Dict[str, int] = {}
    for row in rows:
        for level, count in row["link_levels"].items():
            totals[level] = totals.get(level, 0) + count
    if totals:
        lines.append("**Link precision across all records** (where a hit actually takes the reader): "
                     + ", ".join(f"{count} {LINK_LEVEL_WORD.get(level, level)}"
                                 for level, count in sorted(totals.items(), key=lambda kv: -kv[1])) + ".")
        lines.append("")
    for row in rows:
        lines.append(f"## {STATE_MARK[row['state']]} {row['position']:02d}. {row['name']}")
        lines.append("")
        lines.append(f"- **Folder:** `data_sources/{row['folder']}/`")
        lines.append(f"- **Portal:** {row['url'] or '_none_'}" + (f"  ({row['link']})" if row["link"] else ""))
        lines.append(f"- **Coverage in workbook:** {row['coverage'] or 'not stated'}"
                     + (f" | access: {row['access']}" if row["access"] else "")
                     + (f" | note: {row['note']}" if row["note"] else ""))
        if row["files"]:
            shown = ", ".join(f"`{name}`" for name in row["files"][:6])
            more = f" (+{len(row['files']) - 6} more)" if len(row["files"]) > 6 else ""
            lines.append(f"- **Downloaded:** {len(row['files'])} file(s), {human_bytes(row['bytes'])}: {shown}{more}")
        else:
            lines.append("- **Downloaded:** nothing yet")
        lines.append(f"- **Indexed:** {row['indexed']} indicator-level record(s)"
                     + (" + 1 portal-level record" if row["portal_record"] else ""))
        if row["link_levels"]:
            spelled = ", ".join(f"{count} x {LINK_LEVEL_WORD.get(level, level)}"
                                for level, count in sorted(row["link_levels"].items(), key=lambda kv: -kv[1]))
            if row["unverified_links"]:
                spelled += (f" ({row['unverified_links']} of them not verifiable from here: the target "
                            "portal is a client-rendered app or refuses scripted requests)")
            lines.append(f"- **Link precision:** {spelled}")
        lines.append(f"- **Next step:** {row['next']}")
        lines.append("")
    lines.append("## Candidate sources not in the workbook")
    lines.append("")
    lines.append("Suggested additions, most valuable first. Each would cover something nothing "
                 "currently indexed provides. Nothing here is downloaded yet.")
    lines.append("")
    for name, url, why in CANDIDATES:
        lines.append(f"- **{name}** ({url}): {why}")
    lines.append("")
    CHECKLIST.write_text("\n".join(lines), encoding="utf-8")


def write_workbook(rows: List[Dict[str, Any]], stamp: str) -> None:
    if not WORKBOOK.exists():
        return
    if not WORKBOOK_ORIGINAL.exists():
        WORKBOOK_ORIGINAL.write_bytes(WORKBOOK.read_bytes())

    workbook = openpyxl.load_workbook(WORKBOOK)
    if "Status_GeoDB" in workbook.sheetnames:
        del workbook["Status_GeoDB"]
    sheet = workbook.create_sheet("Status_GeoDB")

    # One language per sheet: the added sheet is English like the rest of the project docs,
    # the original German Tabelle1 is left exactly as it was.
    intro = (f"Integration status for the GeoDB finder (geodb.geolab.soz.uni-bielefeld.de). "
             f"Machine-generated on {stamp} by scripts/build_status_report.py. "
             "Everything on this sheet was added by the assistant (shown in blue); "
             "the original sheet Tabelle1 is untouched.")
    sheet["A1"] = intro
    sheet["A1"].font = AI_BLUE_BOLD
    sheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A1:K1")
    sheet.row_dimensions[1].height = 46

    headers = ["No.", "Data source", "URL", "Link check", "State", "Folder",
               "Downloaded", "Size", "Records indexed", "Link precision", "What is missing / next step"]
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=column, value=title)
        cell.font = AI_BLUE_BOLD
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for offset, row in enumerate(rows, start=3):
        values = [
            row["position"], row["name"], row["url"], row["link"] or "nicht geprüft",
            STATE_WORD[row["state"]], f"data_sources/{row['folder']}/",
            f"{len(row['files'])} file(s)" if row["files"] else "nothing",
            human_bytes(row["bytes"]) if row["bytes"] else "",
            row["indexed"] + row["portal_record"],
            ", ".join(f"{count} x {level}" for level, count in sorted(row["link_levels"].items(), key=lambda kv: -kv[1])) or "portal only",
            row["next"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=column, value=value)
            cell.font = AI_BLUE
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [5, 34, 46, 22, 12, 34, 16, 10, 12, 26, 70]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    sheet.freeze_panes = "A3"
    workbook.save(WORKBOOK)




STATE_DE = {"done": "fertig", "partial": "teilweise", "open": "offen"}
LINK_LEVEL_DE = {
    "indicator": "Indikator", "table": "Tabelle", "statistic": "Statistik",
    "dataset": "Datensatz", "portal": "Portal",
}


def write_deliverable_workbook(rows: List[Dict[str, Any]], stamp: str) -> Optional[Path]:
    """A German, dated copy of the workbook for handing over.

    The working file keeps its English status sheet (the project docs are English), but the
    copy that leaves this repo is read by German colleagues, so it carries exactly one status
    sheet and that sheet is German. The copy is dated in its filename because it is a snapshot:
    the working file keeps moving, a handed-over file must not silently change meaning.
    """
    if not WORKBOOK.exists():
        return None
    DELIVERABLES.mkdir(parents=True, exist_ok=True)
    target = DELIVERABLES / f"Geospatial_Data_Sources_GeoDB_Stand_{stamp}.xlsx"

    workbook = openpyxl.load_workbook(WORKBOOK)
    for name in ("Status_GeoDB", "Stand_GeoDB"):
        if name in workbook.sheetnames:
            del workbook[name]
    sheet = workbook.create_sheet("Stand_GeoDB", 1)

    done = sum(1 for row in rows if row["state"] == "done")
    partial = sum(1 for row in rows if row["state"] == "partial")
    open_ = sum(1 for row in rows if row["state"] == "open")
    indexed = sum(row["indexed"] for row in rows)
    # German thousands separator on the number only: replacing every comma in the sentence
    # turned "26 fertig, 5 teilweise" into "26 fertig. 5 teilweise".
    indexed_de = f"{indexed:,}".replace(",", ".")
    entries_de = f"{indexed + len(rows):,}".replace(",", ".")
    intro = (
        f"Stand der Einbindung in den GeoDB-Finder (geodb.geolab.soz.uni-bielefeld.de), "
        f"Stand {stamp}. {len(rows)} Quellen: {done} fertig, {partial} teilweise, {open_} offen; "
        f"{indexed_de} indexierte Merkmale und Datensätze plus je ein Portaleintrag, "
        f"{entries_de} Einträge insgesamt. "
        "Maschinell erzeugt von scripts/build_status_report.py. Alles auf diesem Blatt wurde "
        "vom Assistenten ergänzt (blau); das Originalblatt Tabelle1 ist unverändert, ergänzt "
        "nur um die neu aufgenommenen Quellen (ebenfalls blau)."
    )
    sheet["A1"] = intro
    sheet["A1"].font = AI_BLUE_BOLD
    sheet["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("A1:K1")
    sheet.row_dimensions[1].height = 60

    headers = ["Nr.", "Datenquelle", "URL", "Portal erreichbar", "Status", "Ordner",
               "Heruntergeladen", "Größe", "Indexierte Einträge", "Linkgenauigkeit",
               "Was fehlt / nächster Schritt"]
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=column, value=title)
        cell.font = AI_BLUE_BOLD
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for offset, row in enumerate(rows, start=3):
        values = [
            row["position"], row["name"], row["url"],
            row["link"] or "nicht geprüft",
            STATE_DE[row["state"]],
            f"data_sources/{row['folder']}/",
            f"{len(row['files'])} Datei(en)" if row["files"] else "nichts",
            human_bytes(row["bytes"]) if row["bytes"] else "",
            row["indexed"] + row["portal_record"],
            ", ".join(f"{count} x {LINK_LEVEL_DE.get(level, level)}"
                      for level, count in sorted(row["link_levels"].items(), key=lambda kv: -kv[1]))
            or "nur Portaleintrag",
            row.get("next_de") or row["next"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=column, value=value)
            cell.font = AI_BLUE
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    legend_row = len(rows) + 4
    for line in [
        "Legende Status: fertig = nichts offen; teilweise = eingebunden, ein vollständigerer "
        "Katalog wäre erreichbar; offen = es fehlt ein Schritt, den nur ein Mensch machen kann.",
        "Legende Linkgenauigkeit: Indikator = der Link öffnet genau das Merkmal; Tabelle = die "
        "Tabelle, die es enthält; Statistik/Datensatz = die Statistik bzw. den Datensatz, der es "
        "enthält; Portal = die Startseite, ab da muss gesucht werden.",
        "Die Spalte 'Indexierte Einträge' zählt Merkmale, Tabellen und Datensätze, nicht Zeilen "
        "der Daten selbst: der Finder indexiert Beschreibungen und verlinkt nach außen.",
    ]:
        cell = sheet.cell(row=legend_row, column=1, value=line)
        cell.font = AI_BLUE
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=11)
        sheet.row_dimensions[legend_row].height = 30
        legend_row += 1

    widths = [5, 34, 46, 22, 12, 34, 16, 10, 14, 26, 78]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    sheet.freeze_panes = "A3"
    workbook.save(target)
    return target


# LaTeX will not break a German compound on its own here, so a name longer than the column
# runs into the status cell ("Arbeitsmarktstatistikfertig"). These two get an explicit break
# point for the printed table only; the workbook keeps the full name.
TABLE_NAME_BREAKS = {
    "Arbeitsmarktstatistik BA": "Arbeitsmarkt- statistik BA",
    "Krankenhausverzeichnis": "Krankenhaus- verzeichnis",
}


def table_name(name: str) -> str:
    for long_form, broken in TABLE_NAME_BREAKS.items():
        if long_form in name:
            return name.replace(long_form, broken)
    return name


def write_progress_table(rows: List[Dict[str, Any]], stamp: str) -> Optional[Path]:
    """A single presentable table of where every source stands, for handing on. Written as
    CSV and rendered to PDF/PNG through tinytable (one canonical table package, notes inside
    the image), so the deliverable is regenerated by the same command that updates the
    checklist and can never drift from it."""
    import csv as _csv
    import shutil
    import subprocess

    DELIVERABLES.mkdir(parents=True, exist_ok=True)
    csv_path = DELIVERABLES / f"{PROGRESS_BASE}.csv"
    header = ["Nr.", "Datenquelle", "Status", "Dat.", "Anz.", "Verlinkung", "Stand / offener Schritt"]

    def plain(text: str) -> str:
        return (text.replace("\u2265", ">=").replace("\u2264", "<=")
                    .replace("\u2019", "'").replace("\u2018", "'"))
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = _csv.writer(handle)
        writer.writerow(header)
        for row in rows:
            levels = sorted(row["link_levels"].items(), key=lambda kv: -kv[1])
            spelled = ", ".join(f"{LINK_LEVEL_DE.get(level, level)} {count}" for level, count in levels) or "Portal 1"
            if row["unverified_links"]:
                spelled += " (ungeprueft)"
            step = row["next_de"]
            writer.writerow([
                row["position"], table_name(plain(row["name"])), STATE_DE[row["state"]],
                len(row["files"]), row["indexed"] + row["portal_record"], spelled,
                plain(row["next_de"]),
            ])

    done = sum(1 for r in rows if r["state"] == "done")
    partial = sum(1 for r in rows if r["state"] == "partial")
    open_count = sum(1 for r in rows if r["state"] == "open")
    # The sheet and this table used to quote different totals for one fact (one counted the
    # portal cards, the other did not), which is how a handoff document ends up with two
    # numbers for the same thing. Both now name the two parts explicitly.
    indicators = sum(r["indexed"] for r in rows)
    portals = sum(r["portal_record"] for r in rows)
    indicators_de = f"{indicators:,}".replace(",", ".")
    entries_total_de = f"{indicators + portals:,}".replace(",", ".")
    summary = (f"{len(rows)} Datenquellen: {done} fertig, {partial} teilweise, {open_count} offen; "
               f"{indicators_de} indexierte Merkmale und Datensätze plus {portals} Portaleinträge, "
               f"{entries_total_de} Einträge insgesamt "
               "(geodb.geolab.soz.uni-bielefeld.de).")

    rscript = shutil.which("Rscript") or "/home/researcher/miniconda3/envs/rstats/bin/Rscript"
    # The rstats env ships libfontconfig.so.1 but the loader does not look there unless told,
    # and TinyTeX then fails with "libfontconfig.so.1: cannot open shared object file" that
    # reads like a broken LaTeX install. Point LD_LIBRARY_PATH at the env that owns Rscript.
    env = dict(os.environ)
    env_lib = Path(rscript).resolve().parent.parent / "lib"
    if env_lib.exists():
        env["LD_LIBRARY_PATH"] = f"{env_lib}:{env.get('LD_LIBRARY_PATH', '')}".rstrip(":")
    try:
        subprocess.run([rscript, str(REPO_ROOT / "scripts" / "render_progress_table.R"),
                        str(csv_path), str(DELIVERABLES / PROGRESS_BASE), stamp, summary],
                       check=True, capture_output=True, timeout=600, cwd=str(REPO_ROOT), env=env)
    except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        detail = getattr(exc, "stderr", b"")
        print(f"[warn] progress table not rendered: {exc} {detail[-400:] if detail else ''}")
        return csv_path
    # The clean folder holds deliverables only; tinytable's LaTeX run drops id*.tex/.log there.
    for debris in DELIVERABLES.glob("id*"):
        if debris.suffix in {".tex", ".log", ".aux", ".out"}:
            debris.unlink(missing_ok=True)
    return DELIVERABLES / f"{PROGRESS_BASE}.pdf"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--no-link-check", action="store_true")
    parser.add_argument("--date", default=date.today().isoformat(),
                        help="stamp written into the outputs (default: today)")
    args = parser.parse_args()

    rows = gather(link_check=not args.no_link_check)
    write_checklist(rows, args.date)
    write_workbook(rows, args.date)
    progress = write_progress_table(rows, args.date)
    deliverable = write_deliverable_workbook(rows, args.date)
    print(json.dumps({
        "sources": len(rows),
        "done": sum(1 for r in rows if r["state"] == "done"),
        "partial": sum(1 for r in rows if r["state"] == "partial"),
        "open": sum(1 for r in rows if r["state"] == "open"),
        "indexed_records": sum(r["indexed"] for r in rows),
        "checklist": str(CHECKLIST),
        "workbook": str(WORKBOOK),
        "deliverable_workbook": str(deliverable) if deliverable else None,
        "progress_table": str(progress) if progress else None,
        "unreachable": [r["name"] for r in rows if r["link"].startswith("unreachable") or r["link"].startswith("HTTP 4") or r["link"].startswith("HTTP 5")],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
